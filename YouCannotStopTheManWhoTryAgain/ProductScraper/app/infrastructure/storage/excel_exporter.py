"""Excel export of scraped products via openpyxl (no embedded images)."""
from __future__ import annotations

import logging
from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.application.ports.exporter import ExporterPort
from app.domain.entities.product import Product
from app.infrastructure.config.settings import OutputSettings


class ExcelExporter(ExporterPort):
    """Writes products.xlsx; images are referenced by local path, never embedded."""

    HEADERS = [
        "Product Name",
        "SKU",
        "Price",
        "Sale Price",
        "Description",
        "Category",
        "Brand",
        "Tags",
        "Colors",
        "Sizes",
        "Availability",
        "Weight",
        "Dimensions",
        "Product URL",
        "Main Image Folder",
        "Main Image File",
        "Gallery Images",
    ]

    def __init__(self, settings: OutputSettings, logger: logging.Logger) -> None:
        self._settings = settings
        self._logger = logger

    def export(self, products: Sequence[Product]) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(self.HEADERS)
        self._style_header(sheet)
        for product in products:
            sheet.append(self._row(product))
        self._autofit(sheet)

        path = self._settings.excel_file
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        self._logger.info("Wrote %d rows to %s", len(products), path)
        return path

    def _row(self, product: Product) -> list[object]:
        return [
            product.name,
            product.sku,
            product.price,
            product.sale_price,
            product.description,
            product.category,
            product.brand,
            ", ".join(product.tags),
            ", ".join(product.colors),
            ", ".join(product.sizes),
            "In stock" if product.availability else "Out of stock",
            product.weight,
            product.dimensions,
            product.url,
            product.local_folder,
            product.main_image,
            ";".join(product.gallery_images),
        ]

    @staticmethod
    def _style_header(sheet: Worksheet) -> None:
        fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        sheet.freeze_panes = "A2"

    @staticmethod
    def _autofit(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            longest = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            sheet.column_dimensions[letter].width = min(max(longest + 2, 10), 80)
